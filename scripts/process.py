#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
DK 变化记录表 —— 解析 / 分析 / 生成网页 引擎

用法:
  处理某一天的两个表:
    python process.py --k 2026-08-14_K.xlsx --d 2026-08-14_D.xlsx
    (日期优先取文件名里的 YYYY-MM-DD，也可用 --date 2026-08-14 指定)

  仅用已有数据重新生成网页:
    python process.py --regen

  自检(合成数据，不碰真实数据):
    python process.py --selftest
"""
import argparse
import datetime
import json
import os
import re
import urllib.request

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CONFIG_PATH = os.path.join(BASE, "config.json")
RECORDS_PATH = os.path.join(BASE, "data", "records.json")
STOCKS_PATH = os.path.join(BASE, "data", "stocks.json")
HTML_PATH = os.path.join(BASE, "index.html")
OBS_PATH = os.path.join(BASE, "data", "obs.json")
BLUECHIPS_PATH = os.path.join(BASE, "data", "bluechips.json")
INDUSTRY_PATH = os.path.join(BASE, "data", "industry_map.json")
APP_TEMPLATE_PATH = os.path.join(BASE, "app_template.html")
OBS_WINDOW = 30            # 观察窗口（自然日）
BENCH_SECID = "1.000985"   # 中证全指（前复权）
BENCH_NAME = "中证全指"


# ---------- 基础工具 ----------
def load_config(path=CONFIG_PATH):
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_json(path, default):
    if not os.path.exists(path):
        return default
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_json(path, obj):
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, ensure_ascii=False, indent=2)


def find_col(headers, candidates):
    low = {h.lower(): i for i, h in enumerate(headers) if h}
    # 1) 精确匹配
    for c in candidates:
        if c in low:
            return low[c]
        if c.lower() in low:
            return low[c.lower()]
    # 2) 子串回退（兼容「涨跌幅(13:25)」这类带后缀表头）
    for c in candidates:
        cl = c.lower()
        if not cl:
            continue
        for h, i in low.items():
            if cl in h:
                return i
    return None


def normalize_code(val, pad):
    s = str(val).strip()
    s = re.sub(r"\.0$", "", s)
    s = s.lstrip("'").strip()
    if pad and s.isdigit():
        s = s.zfill(6)
    return s


def parse_cap(val):
    """把「总市值」列文本（如 '230.24亿' / '4.14亿' / '12.3万'）解析为 亿元 数值。无则返回 None。"""
    if val is None:
        return None
    s = str(val).strip().replace(",", "").replace("，", "")
    if not s:
        return None
    m = re.match(r"^([\d.]+)\s*(千亿|万亿|亿|万)?", s)
    if not m:
        return None
    try:
        num = float(m.group(1))
    except ValueError:
        return None
    unit = m.group(2)
    if unit == "亿":
        return num
    if unit == "万":
        return num / 1e4
    if unit == "千亿":
        return num * 1000.0
    if unit == "万亿":
        return num * 10000.0
    return num / 1e8  # 裸数字视为「元」


def parse_date_from_filename(name, config):
    rx = config.get("file_naming", {}).get("date_in_filename_regex")
    if not rx:
        return None
    m = re.search(rx, os.path.basename(name))
    return m.group(1) if m else None


# ---------- 读取 xlsx ----------
def read_xlsx(path, config, kind=None):
    """读取一个 xlsx。
    kind='K'/'D' 时优先按「选股格式」处理：找 K点/D点 列，值为「符合」的才是信号。
    找不到标记列则按「精简清单格式」处理：整表每行都是信号。
    """
    if load_workbook is None:
        raise SystemExit("[错误] 未安装 openpyxl，无法读取 xlsx。请先 pip install openpyxl。")
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    hr = config.get("sheet", {}).get("header_row", 0)
    headers = [str(c).strip() if c is not None else "" for c in rows[hr]]
    code_i = find_col(headers, config["sheet"]["code_col_candidates"])
    name_i = find_col(headers, config["sheet"].get("name_col_candidates", []))
    metric_i = find_col(headers, config["sheet"].get("metric_col_candidates", []))
    cap_i = find_col(headers, config["sheet"].get("cap_col_candidates", []))
    if code_i is None:
        raise SystemExit(
            "[错误] 在 %s 找不到代码列。\n表头: %s\n候选名: %s"
            % (os.path.basename(path), headers, config["sheet"]["code_col_candidates"])
        )
    # 选股格式：K点/D点 列值为「符合」的才是信号
    marker_i = None
    if kind == "K":
        marker_i = find_col(headers, ["K点"])
    elif kind == "D":
        marker_i = find_col(headers, ["D点"])
    mode = "选股(符合筛选)" if marker_i is not None else "精简清单(整表)"
    out = []
    for r in rows[hr + 1:]:
        if not r or all(c is None for c in r):
            continue
        if marker_i is not None:
            v = r[marker_i]
            if v is None or str(v).strip() != "符合":
                continue
        raw = r[code_i]
        if raw is None:
            continue
        code = normalize_code(raw, config["analysis"]["code_pad_to_6"])
        if not code:
            continue
        name = str(r[name_i]).strip() if (name_i is not None and r[name_i] is not None) else ""
        metric = None
        if metric_i is not None and r[metric_i] is not None:
            try:
                metric = float(r[metric_i])
            except Exception:
                metric = None
        cap = parse_cap(r[cap_i]) if cap_i is not None else None
        out.append({"code": code, "name": name, "metric": metric, "cap": cap})
    print("  [统计] %s  解析模式=%s  命中=%d行  含市值=%d行" % (
        os.path.basename(path), mode, len(out), sum(1 for x in out if x["cap"] is not None)))
    return out


# ---------- 处理一天 ----------
def process_day(k_path, d_path, date, config):
    k = read_xlsx(k_path, config, "K")
    d = read_xlsx(d_path, config, "D")
    # 用户要求排除北交所（92/93/8/4 开头）
    k = [i for i in k if not is_bj(i["code"])]
    d = [i for i in d if not is_bj(i["code"])]
    print("  [统计] %s  K表=%d行  D表=%d行 (已排除北交所)" % (date, len(k), len(d)))

    stocks = load_json(STOCKS_PATH, {})
    for it in k + d:
        if it["code"] and it["name"]:
            stocks[it["code"]] = it["name"]
    save_json(STOCKS_PATH, stocks)

    k_codes = sorted({i["code"] for i in k})
    d_codes = sorted({i["code"] for i in d})
    metrics = {}
    caps = {}
    for i in k + d:
        if i["code"] and i["metric"] is not None:
            metrics[i["code"]] = i["metric"]
        if i["code"] and i["cap"] is not None:
            caps[i["code"]] = i["cap"]

    day = {"date": date, "k": k_codes, "d": d_codes, "metrics": metrics, "caps": caps}

    records = load_json(RECORDS_PATH, {"updated": None, "days": []})
    before = len(records.get("days", []))
    removed = purge_bj(records)
    if removed:
        print("  [北交所] 已从历史数据剔除 %d 个信号" % removed)
    days = [x for x in records.get("days", []) if x["date"] != date]
    days.append(day)
    days.sort(key=lambda x: x["date"])
    records["days"] = days
    records["updated"] = date
    save_json(RECORDS_PATH, records)
    print("  [统计] 历史天数 %d -> %d；本日 K=%d  D=%d" % (before, len(days), len(k_codes), len(d_codes)))
    return day


# ---------- 分析 ----------
def build_analysis(records, stocks, config):
    days = records.get("days", [])
    updated = records.get("updated")

    stocks_set = set(stocks.keys())
    for day in days:
        stocks_set.update(day.get("k", []))
        stocks_set.update(day.get("d", []))

    latest_metrics = days[-1].get("metrics", {}) if days else {}
    latest_caps = days[-1].get("caps", {}) if days else {}
    timelines = {}
    for code in stocks_set:
        tl = []
        for day in days:
            if code in day.get("k", []):
                tl.append({"date": day["date"], "state": "K"})
            if code in day.get("d", []):
                tl.append({"date": day["date"], "state": "D"})
        timelines[code] = tl

    rows = []
    total_reversals = 0
    for code, tl in timelines.items():
        nK = sum(1 for e in tl if e["state"] == "K")
        nD = sum(1 for e in tl if e["state"] == "D")
        reversals = []
        for a, b in zip(tl, tl[1:]):
            # 仅跨天状态变化算「反转」；同日既K又D视为口径冲突，不计入预警
            if a["state"] != b["state"] and a["date"] != b["date"]:
                reversals.append({
                    "from": a["state"], "to": b["state"],
                    "from_date": a["date"], "to_date": b["date"],
                    "same_day": False,
                })
        total_reversals += len(reversals)
        last = tl[-1] if tl else None
        first = tl[0] if tl else None
        cons = 1
        if tl:
            for i in range(len(tl) - 1, 0, -1):
                if tl[i]["state"] == tl[i - 1]["state"]:
                    cons += 1
                else:
                    break
        latest_rev = reversals[-1] if reversals else None
        rows.append({
            "code": code,
            "name": stocks.get(code, ""),
            "metric": latest_metrics.get(code),
            "cap": latest_caps.get(code),
            "nK": nK, "nD": nD,
            "first_date": first["date"] if first else None,
            "last_date": last["date"] if last else None,
            "current": last["state"] if last else None,
            "reversals": reversals,
            "n_reversals": len(reversals),
            "latest_reversal": latest_rev,
            "trailing": cons,
            "timeline": tl,
        })

    def sort_key(r):
        lr = r["latest_reversal"]
        lr_date = lr["to_date"] if lr else "0000-00-00"
        return (lr is not None, lr_date, r["last_date"] or "0000-00-00")

    rows.sort(key=sort_key, reverse=True)

    today = days[-1]["date"] if days else None
    today_k = days[-1].get("k", []) if days else []
    today_d = days[-1].get("d", []) if days else []

    return {
        "updated": updated,
        "today": today,
        "today_k": today_k,
        "today_d": today_d,
        "total_stocks": len(rows),
        "total_reversals": total_reversals,
        "stocks_with_reversal": sum(1 for r in rows if r["reversals"]),
        "rows": rows,
    }


# ---------- 观察池（首次D后30天涨跌幅，验证D参考性）----------
def market_prefix(code):
    """北交所/新三板优先；沪市(6/9开头) -> sh；深市(0/3开头) -> sz。"""
    c = str(code)
    if c[:2] in ("92", "93") or c[0] in ("8", "4"):
        return "bj"
    if c[0] in ("6", "9"):
        return "sh"
    if c[0] in ("0", "3"):
        return "sz"
    return "sh"


def is_bj(code):
    """是否北交所/新三板代码（92/93 开头或 8/4 开头）。用户要求排除北交所股票。"""
    c = str(code)
    return c[:2] in ("92", "93") or c[0] in ("8", "4")


def purge_bj(records):
    """把 records.days 中所有北交所代码剔除（含历史），返回剔除数量。"""
    removed = 0
    for day in records.get("days", []):
        for key in ("k", "d"):
            before = day.get(key, [])
            kept = [c for c in before if not is_bj(c)]
            removed += len(before) - len(kept)
            day[key] = kept
    return removed


# 腾讯 K线多域名轮换：web.ifzq 曾整批 501(接口下线)，裸域名/代理域名被高频限流后也 501。
# 轮换可摊薄单个域名请求量；全部失败后由新浪 quotes.sina.cn 兜底（不受腾讯限流影响）。
TX_KLINE_HOSTS = [
    "https://ifzq.gtimg.cn/appstock/app/fqkline/get",
    "https://proxy.finance.qq.com/ifzqgtimg/appstock/app/fqkline/get",
    "https://web.ifzq.gtimg.cn/appstock/app/fqkline/get",
]
_tx_kline_idx = 0


def _next_tx_host():
    global _tx_kline_idx
    h = TX_KLINE_HOSTS[_tx_kline_idx % len(TX_KLINE_HOSTS)]
    _tx_kline_idx += 1
    return h


def _fetch_sina_kline_full(code, days):
    """新浪日K线兜底，返回 [date,open,close,high,low,vol]（不复权，顺序与腾讯对齐）。"""
    sym = market_prefix(code) + code
    url = ("https://quotes.sina.cn/cn/api/json_v2.php/CN_MarketDataService.getKLineData"
           "?symbol=%s&scale=240&ma=no&datalen=%d" % (sym, max(days, 60)))
    req = urllib.request.Request(url, headers={"Referer": "https://finance.sina.com.cn/"})
    with urllib.request.urlopen(req, timeout=15) as r:
        b = r.read().decode("utf-8", "ignore")
    arr = json.loads(b) if b.strip().startswith("[") else []
    if not arr:
        return None
    out = []
    for x in arr:
        out.append([x["day"], float(x["open"]), float(x["close"]),
                    float(x["high"]), float(x["low"]), float(x.get("volume") or 0)])
    return out


def fetch_kline(code, beg, end, mkt=None):
    """拉前复权日K线（腾讯 3 域名轮换 + 新浪兜底），返回 [{date, close}] 或 None。
    重要：腾讯 fqkline 接口在 beg 接近 end 时会漏掉最近一日；强制把 beg 拉远到 end-200 天
    保证最新收盘日一定包含。早期多余数据由 _close_on_or_before 按 entry_date 过滤。"""
    if mkt is None:
        mkt = market_prefix(code)
    from datetime import date as _date, timedelta
    today = _date.fromisoformat(end)
    beg_use = (today - timedelta(days=200)).isoformat()
    last_err = None
    for _ in range(len(TX_KLINE_HOSTS)):
        host = _next_tx_host()
        url = "%s?param=%s%s,day,%s,%s,200,qfq" % (host, mkt, code, beg_use, end)
        try:
            req = urllib.request.Request(url, headers={"Referer": "https://gu.qq.com/"})
            with urllib.request.urlopen(req, timeout=20) as r:
                b = json.loads(r.read().decode("utf-8", "ignore"))
            node = (b.get("data") or {}).get("%s%s" % (mkt, code)) \
                or (b.get("data") or {}).get(code) or {}
            arr = node.get("qfqday") or node.get("day") or []
            if arr:
                return [{"date": line[0], "close": float(line[2])} for line in arr]
            last_err = "empty"
        except Exception as e:
            last_err = e
    # 腾讯全失败 → 新浪兜底
    try:
        arr = _fetch_sina_kline_full(code, 240)
        if arr:
            return [{"date": x[0], "close": x[2]} for x in arr]
    except Exception as e:
        last_err = e
    print("  [价格] 拉取失败 %s: %s" % (code, last_err))
    return None


def fetch_kline_full(code, days=30, mkt=None):
    """拉日K线（腾讯 3 域名轮换 + 新浪兜底），返回浏览器端 evalPullback 所需的 [date,open,close,high,low,vol] 数组列表。"""
    if mkt is None:
        mkt = market_prefix(code)
    last_err = None
    for _ in range(len(TX_KLINE_HOSTS)):
        host = _next_tx_host()
        url = "%s?param=%s%s,day,,,%d,qfq" % (host, mkt, code, days)
        try:
            req = urllib.request.Request(url, headers={"Referer": "https://gu.qq.com/"})
            with urllib.request.urlopen(req, timeout=20) as r:
                b = json.loads(r.read().decode("utf-8", "ignore"))
            node = (b.get("data") or {}).get("%s%s" % (mkt, code)) \
                or (b.get("data") or {}).get(code) or {}
            arr = node.get("qfqday") or node.get("day") or []
            # 腾讯顺序: date, open, close, high, low, volume; evalPullback 也按此顺序读取
            if arr:
                return [[r[0], float(r[1]), float(r[2]), float(r[3]), float(r[4]), float(r[5])] for r in arr]
            last_err = "empty"
        except Exception as e:
            last_err = e
    # 腾讯全失败 → 新浪兜底
    try:
        arr = _fetch_sina_kline_full(code, days)
        if arr:
            return arr
    except Exception as e:
        last_err = e
    print("  [K线] 拉取失败 %s: %s" % (code, last_err))
    return None


def build_klines_for_pullback(records, days_window=30):
    """为近 N 天出现过 D 点的个股预拉 K 线，避免手机浏览器跨域/网络问题。
    返回 {code: [[date,open,close,high,low,vol], ...]}。"""
    days = records.get("days", [])
    if not days:
        return {}
    from datetime import date as _date
    today = _date.fromisoformat(days[-1]["date"])
    cutoff = (today - datetime.timedelta(days=days_window)).isoformat()
    codes = set()
    for day in days:
        if day["date"] < cutoff:
            continue
        codes.update(day.get("d", []))
    codes = sorted(codes)
    print("  [回踩K线] 需预拉 %d 只近 %d 天 D 点股" % (len(codes), days_window))
    out = {}
    ok = 0
    for code in codes:
        arr = fetch_kline_full(code, days=days_window + 10)
        if arr:
            out[code] = arr
            ok += 1
    print("  [回踩K线] 命中 %d/%d" % (ok, len(codes)))
    return out


# ---------------------------------------------------------------------------
# earnings-radar（财报舆情雷达）数据接入 —— 提供「真实资金面」：
#   行业/概念板块主力净流入、个股净流入/净流出 TOP、ETF 资金、外盘、财报舆情
# 数据由 earnings-radar 服务端每日 08:30 / 22:00 推送至其 latest.json。
# 这里在构建期拉取并烤入种子，作为资金流向 tab 的真实资金层（与 DK 信号面交叉验证）。
# ---------------------------------------------------------------------------
ER_SOURCES = [
    "https://ghproxy.net/https://raw.githubusercontent.com/seonkoo/earnings-radar/main/latest.json",
    "https://seonkoo.github.io/earnings-radar/latest.json",
]


def _er_sec(x):
    """东财板块/个股一条记录 → 紧凑字典（net 单位换算为亿元）。"""
    if not isinstance(x, dict):
        return None
    return {
        "name": x.get("f14") or x.get("name"),
        "code": x.get("f12") or x.get("code"),
        "chg": x.get("f3") if x.get("f3") is not None else x.get("pct"),
        "net": round((x.get("f62") or 0) / 1e8, 2),
    }


def _compact_er(d):
    """把 earnings-radar 的 latest.json 压成资金流向 tab 需要的紧凑结构。"""
    industry = [s for s in (_er_sec(x) for x in d.get("industry", [])) if s]
    concept = [s for s in (_er_sec(x) for x in d.get("concept", [])) if s]
    out = [s for s in (_er_sec(x) for x in d.get("out", [])) if s]
    etf = d.get("etf", {}) or {}
    return {
        "available": True,
        "updated": d.get("updated"),
        "status": d.get("status"),
        "industry_in": sorted(industry, key=lambda z: -z["net"])[:12],
        "industry_out": sorted(out, key=lambda z: z["net"])[:12],
        "concept_in": sorted(concept, key=lambda z: -z["net"])[:12],
        "stocks_in": [_er_sec(x) for x in d.get("stocks", [])[:12] if _er_sec(x)],
        "stocks_out": [_er_sec(x) for x in d.get("outStocks", [])[:8] if _er_sec(x)],
        "etf": {
            "byGroup": {k: round(v / 1e8, 2) for k, v in (etf.get("byGroup") or {}).items()},
            "topIn": [{"name": x.get("name"), "code": x.get("code"),
                       "pct": x.get("pct"), "net": round((x.get("f62") or 0) / 1e8, 2),
                       "group": x.get("group")}
                      for x in (etf.get("topIn") or [])[:8]],
            "topOut": [{"name": x.get("name"), "code": x.get("code"),
                        "pct": x.get("pct"), "net": round((x.get("f62") or 0) / 1e8, 2),
                        "group": x.get("group")}
                       for x in (etf.get("topOut") or [])[:8]],
        },
        "overseas": [{"name": x.get("name"), "pct": x.get("pct")}
                     for x in d.get("overseas", [])],
        "earnings": (d.get("earnings") or {}).get("overview") or {},
    }


# ============================================================
#  真实资金面 · 东财实时行情（不再依赖 earnings-radar 的陈旧快照）
#  - 行业板块 / ETF / 个股主力净流入 全部用 push2delay.eastmoney.com 实时拉取
#  - earnings-radar 仅作增强源（概念/海外/财报），拉不到则优雅降级
# ============================================================
EM_HEADERS = {"User-Agent": "Mozilla/5.0", "Referer": "https://quote.eastmoney.com/"}

# 主流行业板块（东财 BK 代码，已验证有效）；数值实时拉取
EM_INDUSTRY_POOL = [
    "BK0473", "BK0474", "BK0475", "BK0451", "BK0437", "BK0478", "BK0422", "BK1037", "BK1031",
    "BK1036", "BK1038", "BK1039", "BK0448", "BK0730", "BK0738", "BK0739", "BK0480", "BK0735",
    "BK0727", "BK1040", "BK0481", "BK0482", "BK0477", "BK0438", "BK0450", "BK0479", "BK0421",
    "BK0424", "BK0425", "BK0731", "BK0440", "BK0433", "BK0436", "BK0427", "BK0447", "BK0454",
]
# ETF 监控池 (code, market, group)；market: 1=沪 0=深
EM_ETF_POOL = [
    ("588170", 1, "主题·科创半导体"), ("588000", 1, "宽基·科创50"), ("588080", 1, "宽基·科创50"),
    ("510300", 1, "宽基·沪深300"), ("510500", 1, "宽基·中证500"), ("159915", 0, "宽基·创业板"),
    ("512100", 1, "宽基·中证1000"), ("510050", 1, "宽基·上证50"), ("159949", 0, "宽基·创业板50"),
    ("512480", 1, "行业·半导体"), ("588200", 1, "行业·科创芯片"), ("512000", 1, "行业·券商"),
    ("512880", 1, "行业·证券"), ("512660", 1, "行业·军工"), ("515050", 1, "行业·5G"),
    ("516970", 1, "行业·基建"), ("515880", 1, "行业·通信"), ("512760", 1, "行业·芯片"),
    ("159995", 0, "行业·芯片"), ("512690", 1, "行业·酒"), ("512010", 1, "行业·医药"),
    ("515790", 1, "行业·光伏"), ("515030", 1, "行业·新能源车"), ("512800", 1, "行业·银行"),
    ("512200", 1, "行业·房地产"), ("515710", 1, "行业·食品"), ("159928", 0, "行业·消费"),
    ("518880", 1, "商品·黄金"), ("159980", 0, "商品·有色金属"), ("159981", 0, "商品·能源化工"),
    ("513180", 1, "跨境·恒生科技"), ("513120", 1, "跨境·港股创新药"), ("513100", 1, "跨境·纳指"),
    ("159920", 0, "跨境·恒生"), ("513500", 1, "跨境·标普500"),
    ("511360", 1, "货币·短融"), ("511880", 1, "货币·银华日利"),
]


def _em_secid(code, market=None):
    """代码 → 东财 secid。行业/概念 BK 用 90. 前缀；个股/ETF 按市场前缀。"""
    if code.startswith("BK"):
        return "90." + code
    if market is not None:
        return "%d.%s" % (int(market), code)
    lead = code[0]
    return ("0." + code) if lead in "0168" else ("1." + code)


def _ulist_batch(secids, fields="f12,f14,f62,f3,f100,f164,f165,f267,f268", timeout=20):
    url = ("https://push2delay.eastmoney.com/api/qt/ulist.np/get?fltt=2&secids=%s&fields=%s"
           % (secids, fields))
    req = urllib.request.Request(url, headers=EM_HEADERS)
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return (json.loads(r.read().decode("utf-8", "ignore")).get("data") or {}).get("diff") or []


def fetch_em_flow(stock_codes, names, timeout=30):
    """拉取东财实时主力资金（行业板块/ETF/个股），返回与页面兼容的结构。"""
    import datetime
    items = []
    for bk in EM_INDUSTRY_POOL:
        items.append(("90." + bk, "industry", {"code": bk}))
    for code, mkt, grp in EM_ETF_POOL:
        items.append(("%d.%s" % (mkt, code), "etf", {"code": code, "group": grp}))
    for c in stock_codes:
        items.append((_em_secid(c), "stock", {"code": c}))
    got = {}
    for i in range(0, len(items), 50):
        batch = items[i:i + 50]
        secids = ",".join(b[0] for b in batch)
        try:
            diffs = _ulist_batch(secids, timeout=timeout)
            for x in diffs:
                f12 = x.get("f12")
                f14 = x.get("f14")
                if f12 and f14 and f14 != "_":
                    got[f12] = {"name": f14, "net": round((x.get("f62") or 0) / 1e8, 2),
                                "chg": x.get("f3"), "ind": x.get("f100"),
                                "d5": round((x.get("f164") or 0) / 1e8, 2),
                                "d10": round((x.get("f267") or 0) / 1e8, 2)}
        except Exception as e:
            print("  [东财实时] 批次 %d-%d 拉取失败: %s" % (i, i + len(batch), e))
    # 行业板块
    ind = []
    for bk in EM_INDUSTRY_POOL:
        if bk in got:
            v = got[bk]
            ind.append({"name": v["name"], "code": bk, "chg": v["chg"], "net": v["net"]})
    ind_in = sorted(ind, key=lambda z: -z["net"])[:15]
    ind_out = sorted(ind, key=lambda z: z["net"])[:15]
    # ETF
    etf_items = []
    for code, mkt, grp in EM_ETF_POOL:
        if code in got:
            v = got[code]
            etf_items.append({"name": v["name"], "code": code, "pct": v["chg"],
                              "net": v["net"], "group": grp})
    etf_in = sorted(etf_items, key=lambda z: -z["net"])[:10]
    etf_out = sorted(etf_items, key=lambda z: z["net"])[:10]
    byGroup = {}
    for it in etf_items:
        byGroup[it["group"]] = round(byGroup.get(it["group"], 0) + it["net"], 2)
    # 个股（观察池，实时 TOP）
    stk = []
    for c in stock_codes:
        if c in got:
            v = got[c]
            stk.append({"name": names.get(c, c), "code": c, "chg": v["chg"], "net": v["net"]})
    stk_in = sorted(stk, key=lambda z: -z["net"])[:12]
    stk_out = sorted(stk, key=lambda z: z["net"])[:8]
    # 个股原始映射（含所属行业 f100），供「个股→行业」聚合复用，避免重复请求
    raw_stocks = {}
    for c in stock_codes:
        if c in got:
            v = got[c]
            raw_stocks[c] = {"name": names.get(c, v["name"]), "chg": v["chg"],
                             "net": v["net"], "ind": v.get("ind"),
                             "d5": v.get("d5"), "d10": v.get("d10")}
    return {
        "available": True,
        "updated": datetime.datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
        "source": "东财实时(ulist)",
        "industry_in": ind_in, "industry_out": ind_out,
        "etf": {"byGroup": byGroup, "topIn": etf_in, "topOut": etf_out},
        "stocks_in": stk_in, "stocks_out": stk_out,
        "raw_stocks": raw_stocks,
    }


# ============================================================
#  资金全景：个股 → 行业 → 大市主力资金
#  - 大市：上证/深证/创业板 主力·超大单·大单·中单·小单 + 5日/10日势头
#  - 行业：东财行业板块同口径 + 势头判定
#  - 个股→行业：按 f100 二级行业聚合当日 K/D 信号与个股主力净额
#  - 共振：个股聚合行业 × 行业级资金 同向(共振)/反向(背离)
# ============================================================
EM_INDEX_POOL = [
    ("000001", 1, "上证指数"),
    ("399001", 0, "深证成指"),
    ("399006", 0, "创业板指"),
]
EM_FLOW_FIELDS = "f12,f14,f3,f62,f184,f66,f72,f78,f84,f164,f165,f267,f268"


def _yi(v, nd=2):
    """元 → 亿元（None 安全）。"""
    if v is None:
        return None
    try:
        return round(float(v) / 1e8, nd)
    except Exception:
        return None


def momentum_label(t, d5, d10=None):
    """资金势头判定 → (label, tone)。tone: strong_in/in/flat/out/strong_out
    依据：今日主力净额 vs 5日均值 的方向与加速度。"""
    try:
        t = float(t or 0)
    except Exception:
        t = 0.0
    a5 = None if d5 is None else float(d5) / 5.0
    if a5 is None:
        base = "流入" if t > 0 else ("流出" if t < 0 else "持平")
        acc = ""
    elif t > 0 and a5 > 0:
        base = "持续流入"
        acc = "加速" if t > a5 * 1.3 else ("放缓" if t < a5 * 0.7 else "平稳")
    elif t > 0 and a5 <= 0:
        base, acc = "资金回流", "转多"
    elif t < 0 and a5 < 0:
        base = "持续流出"
        acc = "加速" if t < a5 * 1.3 else ("放缓" if t > a5 * 0.7 else "平稳")
    elif t < 0 and a5 >= 0:
        base, acc = "高位流出", "转空"
    else:
        base, acc = "窄幅波动", ""
    label = base + (("·" + acc) if acc else "")
    if t > 0 and (a5 or 0) > 0:
        tone = "strong_in"
    elif t > 0:
        tone = "in"
    elif t < 0 and (a5 or 0) < 0:
        tone = "strong_out"
    elif t < 0:
        tone = "out"
    else:
        tone = "flat"
    return label, tone


def fetch_industry_boards(pages=5, page_size=100, timeout=25):
    """拉取东财全部行业板块资金流（clist, fs=m:90+t:2），按主力净额降序。
    板块名与个股 f100 二级行业口径一致，可直接对照。"""
    import urllib.request
    out = []
    seen = set()
    for pn in range(1, pages + 1):
        url = ("https://push2delay.eastmoney.com/api/qt/clist/get?pn=%d&pz=%d&po=1&np=1"
               "&fltt=2&invt=2&fid=f62&fs=m:90+t:2&fields=%s"
               % (pn, page_size, EM_FLOW_FIELDS))
        try:
            req = urllib.request.Request(url, headers=EM_HEADERS)
            with urllib.request.urlopen(req, timeout=timeout) as r:
                d = json.loads(r.read().decode("utf-8", "ignore"))
            diff = ((d.get("data") or {}).get("diff")) or []
            for x in diff:
                code = x.get("f12")
                if not code or code in seen:
                    continue
                seen.add(code)
                main = _yi(x.get("f62"))
                if main is None:
                    continue
                d5, d10 = _yi(x.get("f164")), _yi(x.get("f267"))
                lab, tone = momentum_label(main, d5, d10)
                out.append({"code": code, "name": x.get("f14"), "chg": x.get("f3"),
                            "main": main, "ratio": x.get("f184"),
                            "super": _yi(x.get("f66")), "big": _yi(x.get("f72")),
                            "mid": _yi(x.get("f78")), "small": _yi(x.get("f84")),
                            "d5": d5, "d5r": x.get("f165"),
                            "d10": d10, "d10r": x.get("f268"),
                            "mom": lab, "tone": tone})
            if len(diff) < page_size:
                break
        except Exception as e:
            print("  [行业板块] 第%d页拉取失败: %s" % (pn, e))
            break
    out.sort(key=lambda z: -(z["main"] or 0))
    return out


def fetch_market_panorama(raw_stocks, records=None, obs=None, timeout=25):
    """构建「个股→行业→大市」资金全景。raw_stocks 来自 fetch_em_flow。"""
    import datetime
    out = {"available": False, "updated": None, "source": "东财实时(ulist)",
           "indices": [], "summary": None, "industries": [],
           "stock_industry": [], "verdict": None}
    try:
        # ---- 1) 大市指数（一次批量）
        secids = ",".join(["%d.%s" % (mkt, c) for c, mkt, _ in EM_INDEX_POOL])
        got = {}
        try:
            for x in _ulist_batch(secids, fields=EM_FLOW_FIELDS, timeout=timeout):
                f12 = x.get("f12")
                if f12 and x.get("f14") not in (None, "_"):
                    got[f12] = x
        except Exception as e:
            print("  [资金全景] 大市指数拉取失败: %s" % e)

        def row_of(x):
            main = _yi(x.get("f62"))
            d5 = _yi(x.get("f164"))
            d10 = _yi(x.get("f267"))
            label, tone = momentum_label(main, d5, d10)
            return {"name": x.get("f14"), "chg": x.get("f3"),
                    "main": main, "ratio": x.get("f184"),
                    "super": _yi(x.get("f66")), "big": _yi(x.get("f72")),
                    "mid": _yi(x.get("f78")), "small": _yi(x.get("f84")),
                    "d5": d5, "d5r": x.get("f165"),
                    "d10": d10, "d10r": x.get("f268"),
                    "mom": label, "tone": tone}

        indices = []
        for code, mkt, disp in EM_INDEX_POOL:
            if code in got:
                r = row_of(got[code])
                r["code"] = code
                r["label"] = disp
                indices.append(r)

        # 行业板块资金：东财全量行业板块口径（分页拉取，按主力净额降序）
        industries = fetch_industry_boards(timeout=timeout)

        # ---- 2) 沪深两市合计（上证 + 深证）
        summary = None
        if indices:
            def sget(key):
                vals = [i[key] for i in indices if i.get(key) is not None]
                return round(sum(vals), 2) if vals else None
            ss, mm = sget("main"), sget("mid")
            sm, bb = sget("small"), sget("big")
            sup = sget("super")
            d5, d10 = sget("d5"), sget("d10")
            lab, tone = momentum_label(ss, d5, d10)
            summary = {"main": ss, "super": sup, "big": bb, "mid": mm, "small": sm,
                       "d5": d5, "d10": d10, "mom": lab, "tone": tone}

        # ---- 3) 个股 → 行业聚合（当日 K/D 信号 + 个股主力净额）
        days = (records or {}).get("days", []) or []
        today = days[-1] if days else {}
        kset = set(today.get("k", []) or [])
        dset = set(today.get("d", []) or [])
        agg = {}
        for code, v in (raw_stocks or {}).items():
            ind = v.get("ind")
            if not ind:
                continue
            a = agg.setdefault(ind, {"industry": ind, "n": 0, "nK": 0, "nD": 0,
                                     "chg_sum": 0.0, "chg_n": 0, "net": 0.0,
                                     "d5": 0.0, "d10": 0.0})
            a["n"] += 1
            if code in kset:
                a["nK"] += 1
            if code in dset:
                a["nD"] += 1
            if isinstance(v.get("chg"), (int, float)):
                a["chg_sum"] += float(v["chg"])
                a["chg_n"] += 1
            if isinstance(v.get("net"), (int, float)):
                a["net"] += float(v["net"])
            if isinstance(v.get("d5"), (int, float)):
                a["d5"] += float(v["d5"])
            if isinstance(v.get("d10"), (int, float)):
                a["d10"] += float(v["d10"])
        stock_industry = []
        for a in agg.values():
            a["net"] = round(a["net"], 2)
            a["d5"] = round(a["d5"], 2)
            a["d10"] = round(a["d10"], 2)
            a["net_signal"] = a["nK"] - a["nD"]          # >0 卖出占优
            a["avg_chg"] = round(a["chg_sum"] / a["chg_n"], 2) if a["chg_n"] else None
            a.pop("chg_sum", None)
            a.pop("chg_n", None)
            # 势头：由个股汇总的 今日 / 5日 / 10日 主力净额推算
            lab, tone = momentum_label(a["net"], a["d5"], a["d10"])
            a["mom"], a["tone"] = lab, tone
            stock_industry.append(a)
        stock_industry.sort(key=lambda z: -abs(z["net"] or 0))

        # ---- 4) 共振：个股聚合资金（自下而上）× 该行业 K/D 信号方向
        for a in stock_industry:
            bull_sig = a["net_signal"] < 0               # D 多 = 买入占优
            bull_money = (a["net"] or 0) > 0
            if bull_sig and bull_money:
                a["reso"] = "共振看多"
            elif (not bull_sig) and (not bull_money):
                a["reso"] = "共振看空"
            elif bull_sig and not bull_money:
                a["reso"] = "背离·信号多资金出"
            else:
                a["reso"] = "背离·资金进信号弱"

        # ---- 5) 大市研判
        verdict = build_market_verdict(summary, industries, stock_industry,
                                       kset, dset, obs)

        out.update({"available": True,
                    "updated": datetime.datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
                    "indices": indices, "summary": summary,
                    "industries": industries, "stock_industry": stock_industry,
                    "verdict": verdict})
        n_ind = len(industries)
        print("  [资金全景] 指数%d 行业%d 个股行业聚合%d 共振已计算"
              % (len(indices), n_ind, len(stock_industry)))
    except Exception as e:
        print("  [资金全景] 构建失败: %s" % e)
    return out


def build_market_verdict(summary, industries, stock_industry, kset, dset, obs):
    """综合「大市主力资金 + 行业宽度 + 信号面 + 观察池实况」给出研判。"""
    main = (summary or {}).get("main") or 0
    # 方向与强度
    if main > 100:
        direction, strength = "大幅净流入", "强"
    elif main > 20:
        direction, strength = "净流入", "偏强"
    elif main > -20:
        direction, strength = "基本平衡", "中性"
    elif main > -100:
        direction, strength = "净流出", "偏弱"
    else:
        direction, strength = "大幅净流出", "弱"
    # 行业宽度
    tot = len(industries) or 1
    in_n = sum(1 for i in industries if (i.get("main") or 0) > 0)
    breadth = round(in_n * 100.0 / tot, 1)
    # 信号面
    nk, nd = len(kset), len(dset)
    sig = "买入占优" if nd > nk else ("卖出占优" if nk > nd else "多空均衡")
    # 观察池实况
    sm = (obs or {}).get("summary") or {}
    win = sm.get("win_rate")
    avg = sm.get("avg_ret")
    # 共振统计
    reso_bull = sum(1 for a in stock_industry if a.get("reso") == "共振看多")
    reso_bear = sum(1 for a in stock_industry if a.get("reso") == "共振看空")

    # 结论（严格以价格动作为准，不臆造）
    parts = []
    parts.append("沪深两市主力资金%s（%+.0f亿，强度%s）" % (direction, main, strength))
    parts.append("行业净流入宽度 %.0f%%（%d/%d）" % (breadth, in_n, tot))
    parts.append("今日信号 K=%d D=%d，%s" % (nk, nd, sig))
    if win is not None:
        parts.append("观察池胜率 %.1f%%、均收益 %+.2f%%" % (win, avg or 0))
    # 情绪判定：以观察池价格动作为准
    if avg is not None and avg < 0 and (win or 0) < 50:
        mood = "仍在退潮"
    elif avg is not None and avg > 0 and (win or 0) >= 50:
        mood = "情绪已修复"
    else:
        mood = "尚未修复，局部企稳"
    text = "；".join(parts) + "。综合研判：" + mood + "。"
    return {"direction": direction, "strength": strength, "main": round(main, 2),
            "breadth": breadth, "in_n": in_n, "tot": tot,
            "nk": nk, "nd": nd, "sig": sig,
            "reso_bull": reso_bull, "reso_bear": reso_bear,
            "win_rate": win, "avg_ret": avg, "mood": mood, "text": text}


def build_realtime_er(stocks, config):
    """实时资金面 = 东财实时核心 + earnings-radar 增强（概念/海外/财报，best-effort）。"""
    names = stocks if isinstance(stocks, dict) else {}
    em = fetch_em_flow(list(names.keys()), names)
    er_raw = fetch_earnings_radar()
    if er_raw:
        em["concept_in"] = er_raw.get("concept_in", []) or []
        em["overseas"] = er_raw.get("overseas", []) or []
        em["earnings"] = er_raw.get("earnings", {}) or {}
        em["er_source"] = "earnings-radar"
        print("  [资金面] 东财实时(行业/ETF/个股) + earnings-radar 增强(概念/海外/财报) 已并入")
    else:
        em["concept_in"] = []
        em["overseas"] = []
        em["earnings"] = {}
        em["er_source"] = "unavailable"
        print("  [资金面] 东财实时(行业/ETF/个股)；earnings-radar 不可达，概念/海外/财报降级隐藏")
    return em


def fetch_earnings_radar(timeout=20):
    """拉取 earnings-radar 最新快照；失败返回 None（UI 降级为 unavailable）。"""
    last_err = ""
    for url in ER_SOURCES:
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
            with urllib.request.urlopen(req, timeout=timeout) as r:
                data = json.loads(r.read().decode("utf-8", "ignore"))
            comp = _compact_er(data)
            print("  [earnings-radar] 已接入 更新于 %s (板块%d/概念%d/个股%d)"
                  % (comp["updated"], len(comp["industry_in"]), len(comp["concept_in"]),
                     len(comp["stocks_in"])))
            return comp
        except Exception as e:
            last_err = "%s: %s" % (url, e)
            print("  [earnings-radar] 拉取失败 %s" % last_err)
    return None


def median_py(arr):
    if not arr:
        return None
    a = sorted(arr)
    m = len(a) // 2
    return a[m] if len(a) % 2 else (a[m - 1] + a[m]) / 2


def _close_on_or_before(klines, date):
    best = None
    for k in klines:
        if k["date"] <= date:
            best = k
        else:
            break
    return best


def _gtimg_secid(code):
    """腾讯实时报价 secid 前缀：沪市 sh / 深市 sz / 北交 bj。"""
    if code.startswith("6"):
        return "sh" + code
    if code.startswith(("8", "4")):
        return "bj" + code
    return "sz" + code


def fetch_mktcap_map(codes):
    """批量取总市值(亿元)。腾讯实时报价 [45]=总市值(亿)，一次可传多只。"""
    out = {}
    codes = [c for c in codes if c]
    for i in range(0, len(codes), 50):
        batch = codes[i:i + 50]
        url = "https://qt.gtimg.cn/q=" + ",".join(_gtimg_secid(c) for c in batch)
        try:
            req = urllib.request.Request(url, headers={"Referer": "https://gu.qq.com/"})
            with urllib.request.urlopen(req, timeout=15) as r:
                txt = r.read().decode("gbk", "ignore")
            for line in txt.split(";"):
                if "=" not in line:
                    continue
                key = line.split("=")[0].strip().replace("v_", "")
                seg = line.split('"', 1)
                if len(seg) < 2:
                    continue
                part = seg[1].rsplit('"', 1)[0]
                f = part.split("~")
                code = key[2:] if key[:2] in ("sh", "sz", "bj") else key
                if len(f) > 45 and f[45]:
                    try:
                        out[code] = float(f[45])
                    except Exception:
                        pass
        except Exception as e:
            print("  [市值] 批量拉取失败: %s" % e)
    return out


def fetch_inflow_map(codes):
    """批量取主力净流入(元)。东方财富数据中心 PRIME_INFLOW（沙箱可达，该主机未被封）。
    该接口仅支持单条件过滤，故单码逐查 + 并发（8 线程）。失败则该股为 None。"""
    import concurrent.futures
    import urllib.parse
    def one(code):
        flt = '(SECURITY_CODE="%s")' % code
        url = ("https://datacenter-web.eastmoney.com/api/data/v1/get?reportName=RPT_DMSK_TS_STOCKNEW"
               "&columns=SECURITY_CODE,PRIME_INFLOW&pageSize=1&source=WEB&client=WEB&filter="
               + urllib.parse.quote(flt))
        try:
            req = urllib.request.Request(url, headers={
                "User-Agent": "Mozilla/5.0", "Referer": "https://data.eastmoney.com/"})
            with urllib.request.urlopen(req, timeout=12) as r:
                d = json.loads(r.read().decode("utf-8", "ignore"))
            rows = (d.get("result") or {}).get("data") or []
            if rows:
                return code, rows[0].get("PRIME_INFLOW")
        except Exception:
            pass
        return code, None
    out = {}
    with concurrent.futures.ThreadPoolExecutor(max_workers=8) as ex:
        for c, v in ex.map(one, codes):
            out[c] = v
    return out



def build_obs(records, config):
    """构建观察池：每只首次出现 D 的个股，以其 D日收盘价为起点，
    统计到「满30天 或 首次出现 K点」为止的累计涨幅，并对比基准（中证全指）。
    蓝筹股额外标记。退出后冻结。"""
    days = records.get("days", [])
    if not days:
        return {"updated": records.get("updated"), "benchmark": BENCH_NAME,
                "stocks": [], "summary": {}}
    today = datetime.date.today()
    bluechips = set(load_json(BLUECHIPS_PATH, {}).keys())
    # 每只股票的「首次 D 日」
    first_d = {}
    for day in days:
        for c in day.get("d", []):
            if c not in first_d:
                first_d[c] = day["date"]
    # 每只股票「首次 D 之后」第一次出现 K 的日期（用于退出判定）
    first_k_after = {}
    for c, dd in first_d.items():
        for day in days:
            if day["date"] > dd and c in day.get("k", []):
                first_k_after[c] = day["date"]
                break
    names = load_json(STOCKS_PATH, {})
    cache = {s["code"]: s for s in load_json(OBS_PATH, {}).get("stocks", [])}
    bench_cache = None
    result = []
    for code, entry_date in first_d.items():
        ed = datetime.date.fromisoformat(entry_date)
        # 退出日 = min(首次D+30天, 首次K日)
        exit_date = ed + datetime.timedelta(days=OBS_WINDOW)
        exit_reason = "满30天"
        if code in first_k_after:
            fk = datetime.date.fromisoformat(first_k_after[code])
            if fk < exit_date:
                exit_date = fk
                exit_reason = "出现K点"
        exited = today >= exit_date
        stat_end = exit_date if exited else today
        cached = cache.get(code)
        if cached and cached.get("exited"):
            c2 = dict(cached)
            c2.setdefault("day_chg", None)
            result.append(c2)  # 已流出则冻结，不再拉取
            continue
        klines = fetch_kline(code, entry_date, stat_end.isoformat())
        if klines is None:
            ec = cached.get("entry_close") if cached else None
            ret = cached.get("ret") if cached else None
            br = cached.get("bench_ret") if cached else None
            result.append({
                "code": code, "name": names.get(code, ""), "entry_date": entry_date,
                "entry_close": ec, "ret": ret, "bench_ret": br,
                "day_chg": cached.get("day_chg") if cached else None,
                "excess": (ret - br) if (ret is not None and br is not None) else None,
                "holding_days": (today - ed).days, "exited": exited,
                "exit_reason": exit_reason,
                "status": "待同步" if ret is None else ("已流出" if exited else "活跃"),
                "is_blue": code in bluechips, "updated": today.isoformat(),
            })
            continue
        ek = _close_on_or_before(klines, entry_date) or klines[0]
        entry_close = ek["close"]
        ret = klines[-1]["close"] / entry_close - 1
        # 当日（最新交易日）涨跌幅：最后一根相对前一日的比例
        day_chg = (klines[-1]["close"] / klines[-2]["close"] - 1) if len(klines) >= 2 else None
        if bench_cache is None:
            bench_cache = fetch_kline("000985", entry_date, stat_end.isoformat(), mkt="sh")
        bench_ret = None
        if bench_cache:
            b0 = _close_on_or_before(bench_cache, entry_date)
            if b0:
                bench_ret = bench_cache[-1]["close"] / b0["close"] - 1
        result.append({
            "code": code, "name": names.get(code, ""), "entry_date": entry_date,
            "entry_close": entry_close, "ret": ret, "bench_ret": bench_ret,
            "day_chg": day_chg,
            "excess": (ret - bench_ret) if bench_ret is not None else None,
            "holding_days": (today - ed).days, "exited": exited,
            "exit_reason": exit_reason,
            "status": "已流出" if exited else "活跃",
            "is_blue": code in bluechips, "updated": today.isoformat(),
        })
    # 附加：总市值(亿) 与 资金净流入/主力净流入(亿元)
    all_codes = [x["code"] for x in result]
    mcap_map = fetch_mktcap_map(all_codes)
    inflow_map = fetch_inflow_map(all_codes)
    ok_mcap = sum(1 for v in mcap_map.values() if v is not None)
    ok_inflow = sum(1 for v in inflow_map.values() if isinstance(v, (int, float)))
    print("  [观察池] 总市值命中=%d/%d  资金净流入命中=%d/%d"
          % (ok_mcap, len(all_codes), ok_inflow, len(all_codes)))
    for x in result:
        x["mktcap"] = mcap_map.get(x["code"])                       # 亿元
        ni = inflow_map.get(x["code"])
        x["net_inflow"] = (ni / 1e8) if isinstance(ni, (int, float)) else None  # 元→亿元
    # 排序：活跃在前，其次已流出，最后待同步；各组内按累计涨幅降序
    order = {"活跃": 0, "已流出": 1, "待同步": 2}
    result.sort(key=lambda x: (order.get(x["status"], 3),
                               -(x["ret"] if x["ret"] is not None else -1e9)))
    obs = {"updated": today.isoformat(), "benchmark": BENCH_NAME, "stocks": result}
    # 汇总统计
    # 胜率分母：排除「待同步」，并排除「持有0天」（当天才出现D、尚无可评估收益）的同日项
    valid = [x for x in result if x["ret"] is not None]
    measurable = [x for x in valid if x["holding_days"] >= 1]
    up = [x for x in measurable if x["ret"] > 0]
    down = [x for x in measurable if x["ret"] < 0]
    flat = [x for x in measurable if x["ret"] == 0]
    rets = [x["ret"] for x in measurable]
    obs["summary"] = {
        "total_observed": len(result),
        "valid": len(valid),
        "measurable": len(measurable),
        "up": len(up), "down": len(down), "flat": len(flat),
        "win_rate": round(len(up) / len(measurable) * 100, 1) if measurable else None,
        "pending": sum(1 for x in result if x["status"] == "待同步"),
        "active": sum(1 for x in result if x["status"] == "活跃"),
        "exited": sum(1 for x in result if x["exited"]),
        "exited_30": sum(1 for x in result if x["exited"] and x["exit_reason"] == "满30天"),
        "exited_k": sum(1 for x in result if x["exited"] and x["exit_reason"] == "出现K点"),
        "avg_ret": round(sum(rets) / len(rets) * 100, 2) if rets else None,
        "median_ret": round(median_py(rets) * 100, 2) if rets else None,
        "blue": sum(1 for x in result if x["is_blue"]),
        "blue_win": sum(1 for x in up if x["is_blue"]),
    }
    save_json(OBS_PATH, obs)
    print("  [观察池] 首次D=%d  活跃=%d  已流出=%d(满30天=%d,出K=%d)  待同步=%d"
          % (len(result), obs["summary"]["active"], obs["summary"]["exited"],
             obs["summary"]["exited_30"], obs["summary"]["exited_k"], obs["summary"]["pending"]))
    print("  [汇总] 有效=%d  上涨=%d  下跌=%d  胜率=%s%%  中位涨幅=%s%%"
          % (obs["summary"]["valid"], obs["summary"]["up"], obs["summary"]["down"],
             obs["summary"]["win_rate"], obs["summary"]["median_ret"]))
    return obs


# ---------- 生成网页 ----------
HTML_TEMPLATE = """<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>DK 变化记录表</title>
<style>
  :root{
    --bg:#f6f7f9; --card:#ffffff; --fg:#1c2024; --muted:#6b7280;
    --line:#e5e7eb; --k:#dc2626; --d:#16a34a; --warn:#b45309; --warnbg:#fffbeb;
    --accent:#2563eb;
  }
  @media (prefers-color-scheme: dark){
    :root{
      --bg:#0f1115; --card:#171a21; --fg:#e8eaed; --muted:#9aa3af;
      --line:#262b33; --k:#f87171; --d:#4ade80; --warn:#fbbf24; --warnbg:#2a2113;
      --accent:#60a5fa;
    }
  }
  *{box-sizing:border-box}
  body{margin:0;background:var(--bg);color:var(--fg);
    font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC","Microsoft YaHei",sans-serif;
    line-height:1.5;padding:16px;max-width:960px;margin:0 auto}
  h1{font-size:20px;margin:0 0 4px}
  .sub{color:var(--muted);font-size:13px;margin-bottom:16px}
  .cards{display:grid;grid-template-columns:repeat(4,1fr);gap:10px;margin-bottom:18px}
  @media(max-width:640px){.cards{grid-template-columns:repeat(2,1fr)}}
  .card{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:12px}
  .card .n{font-size:22px;font-weight:700}
  .card .l{font-size:12px;color:var(--muted);margin-top:2px}
  .card.k .n{color:var(--k)} .card.d .n{color:var(--d)}
  .warnbox{background:var(--warnbg);border:1px solid var(--warn);border-radius:12px;
    padding:12px 14px;margin-bottom:18px}
  .warnbox h2{font-size:15px;margin:0 0 8px;color:var(--warn)}
  .witem{display:flex;flex-wrap:wrap;gap:8px;align-items:center;padding:7px 0;
    border-top:1px dashed var(--line);font-size:14px}
  .witem:first-of-type{border-top:none}
  .badge{font-size:11px;padding:1px 7px;border-radius:999px;font-weight:700}
  .bK{background:color-mix(in srgb,var(--k) 18%,transparent);color:var(--k)}
  .bD{background:color-mix(in srgb,var(--d) 18%,transparent);color:var(--d)}
  .bSame{background:var(--warn);color:#fff}
  .arrow{color:var(--muted);font-weight:700}
  .controls{display:flex;flex-wrap:wrap;gap:8px;margin-bottom:12px}
  .controls input{flex:1;min-width:160px;padding:8px 10px;border-radius:9px;
    border:1px solid var(--line);background:var(--card);color:var(--fg);font-size:14px}
  .controls button{padding:8px 12px;border-radius:9px;border:1px solid var(--line);
    background:var(--card);color:var(--fg);font-size:13px;cursor:pointer}
  .controls button.on{background:var(--accent);color:#fff;border-color:var(--accent)}
  table{width:100%;border-collapse:collapse;font-size:13px;background:var(--card);
    border:1px solid var(--line);border-radius:12px;overflow:hidden}
  th,td{padding:8px 10px;text-align:left;border-bottom:1px solid var(--line);white-space:nowrap}
  th{color:var(--muted);font-weight:600;position:sticky;top:0;background:var(--card)}
  tr:last-child td{border-bottom:none}
  td.wrap{white-space:normal}
  .sK{color:var(--k);font-weight:700} .sD{color:var(--d);font-weight:700}
  .tl{color:var(--muted);font-size:12px}
  .empty{color:var(--muted);padding:20px;text-align:center}
  a{color:var(--accent)}
</style>
</head>
<body>
<h1>DK 变化记录表</h1>
<div class="sub">最后更新：__UPDATED__ ｜ K→D / D→K 反转自动置顶预警</div>

<div class="cards" id="cards"></div>
<div class="warnbox" id="warnbox">
  <h2>⚠ 反转预警（重点提示）</h2>
  <div id="warns"></div>
</div>

<div class="controls">
  <input id="q" placeholder="搜索代码 / 名称…" oninput="render()">
  <button id="f_all" class="on" onclick="setF('all',this)">全部</button>
  <button id="f_rev" onclick="setF('rev',this)">仅反转</button>
  <button id="f_k" onclick="setF('k',this)">当前K</button>
  <button id="f_d" onclick="setF('d',this)">当前D</button>
  <select id="sort" onchange="render()" style="padding:8px 10px;border-radius:9px;border:1px solid var(--line);background:var(--card);color:var(--fg);font-size:13px">
    <option value="cap_desc">市值↓ 大→小</option>
    <option value="cap_asc">市值↑ 小→大</option>
    <option value="rev">反转优先</option>
    <option value="code">代码</option>
  </select>
</div>

<table>
  <thead><tr>
    <th>代码</th><th>名称</th><th>市值</th><th>当前</th><th>涨幅%</th><th>首现</th><th>末次</th>
    <th>K</th><th>D</th><th>反转</th><th>连续</th><th>状态时间线</th>
  </tr></thead>
  <tbody id="tbody"></tbody>
</table>

<script>
const DATA = __DATA__;
let FILTER = "all";

function setF(v,el){
  FILTER=v;
  document.querySelectorAll('.controls button').forEach(b=>b.classList.remove('on'));
  el.classList.add('on');
  render();
}
function esc(s){return (s==null?'':String(s));}
function fmtMetric(m){
  if(m==null) return '-';
  const cls = m>0?'sK':(m<0?'sD':'');
  const sign = m>0?'+':'';
  return '<span class="'+cls+'">'+sign+(+m).toFixed(2)+'</span>';
}
function fmtCap(c){ if(c==null) return '-'; if(c>=10000) return (c/10000).toFixed(2)+'万亿'; return c.toFixed(2)+'亿'; }
function sortRows(rows){
  const v=document.getElementById('sort').value; const a=[...rows];
  if(v==='cap_desc') a.sort((x,y)=>(y.cap||0)-(x.cap||0));
  else if(v==='cap_asc') a.sort((x,y)=>(x.cap||0)-(y.cap||0));
  else if(v==='rev') a.sort((x,y)=>((y.latest_reversal?1:0)-(x.latest_reversal?1:0))||((y.last_date||'')<(x.last_date||'')?1:-1));
  else a.sort((x,y)=>x.code<y.code?-1:(x.code>y.code?1:0));
  return a;
}
function stateBadge(s){
  if(s==='K') return '<span class="badge bK">K</span>';
  if(s==='D') return '<span class="badge bD">D</span>';
  return '';
}
function tlText(tl){
  return tl.map(e=>{
    const m=e.date.slice(5);
    return (e.state==='K'?'<span class=sK>K</span>':'<span class=sD>D</span>')+'<span class=tl>'+m+'</span>';
  }).join(' → ');
}
function render(){
  // cards
  const c=document.getElementById('cards');
  c.innerHTML=
    card(DATA.total_stocks,'跟踪个股')+
    card(DATA.today_k.length,'今日 K', 'k')+
    card(DATA.today_d.length,'今日 D', 'd')+
    card(DATA.stocks_with_reversal,'反转个股');
  // warns
  const warns=DATA.rows.filter(r=>r.latest_reversal);
  const w=document.getElementById('warns');
  if(!warns.length){w.innerHTML='<div class="empty">暂无反转记录</div>';}
  else{
    w.innerHTML=warns.map(r=>{
      const lr=r.latest_reversal;
      const sd=lr.same_day?' <span class="badge bSame">同日</span>':'';
      const dir=(lr.from==='K'?'K→D':'D→K');
      return '<div class="witem">'+stateBadge(lr.from)+'<span class="arrow">→</span>'+
        stateBadge(lr.to)+sd+' '+
        '<b>'+esc(r.code)+'</b> '+esc(r.name)+' '+
        '<span class="tl">'+lr.from_date+' → '+lr.to_date+'</span>'+
        '<span class="tl">（累计反转 '+r.n_reversals+' 次）</span></div>';
    }).join('');
  }
  // table
  const q=document.getElementById('q').value.trim().toLowerCase();
  const tb=document.getElementById('tbody');
  const rows=DATA.rows.filter(r=>{
    if(FILTER==='rev' && r.reversals.length===0) return false;
    if(FILTER==='k' && r.current!=='K') return false;
    if(FILTER==='d' && r.current!=='D') return false;
    if(q && !(r.code.toLowerCase().includes(q)||(r.name||'').toLowerCase().includes(q))) return false;
    return true;
  });
  if(!rows.length){tb.innerHTML='<tr><td colspan="12" class="empty">无匹配</td></tr>';return;}
  tb.innerHTML=sortRows(rows).map(r=>{
    const cur=r.current==='K'?'<span class=sK>K</span>':(r.current==='D'?'<span class=sD>D</span>':'-');
    const rev=r.latest_reversal?('<b style="color:var(--warn)">'+(r.latest_reversal.from==='K'?'K→D':'D→K')+'</b>'):'-';
    return '<tr><td>'+esc(r.code)+'</td><td>'+esc(r.name)+'</td><td>'+fmtCap(r.cap)+'</td><td>'+cur+'</td><td>'+fmtMetric(r.metric)+'</td>'+
      '<td class="tl">'+esc(r.first_date)+'</td><td class="tl">'+esc(r.last_date)+'</td>'+
      '<td>'+r.nK+'</td><td>'+r.nD+'</td><td>'+rev+'</td><td>'+r.trailing+'</td>'+
      '<td class="wrap tl">'+tlText(r.timeline)+'</td></tr>';
  }).join('');
}
function card(n,l,cls){return '<div class="card '+(cls||'')+'"><div class="n">'+n+'</div><div class="l">'+l+'</div></div>';}
render();
</script>
</body>
</html>
"""


def render_html(analysis, config):
    return HTML_TEMPLATE.replace("__UPDATED__", analysis.get("updated") or "—").replace(
        "__DATA__", json.dumps(analysis, ensure_ascii=False)
    )


def write_html(html):
    with open(HTML_PATH, "w", encoding="utf-8") as f:
        f.write(html)
    print("  [生成] %s (%d 字节)" % (HTML_PATH, len(html)))


def style_of(code, bluechips):
    """把代码归类到风格/板块：蓝筹 > 科技成长(科创/创业) > 北交 > 主板。"""
    if code in bluechips:
        return "蓝筹"
    if code.startswith("688") or code.startswith("300") or code.startswith("301"):
        return "科技成长"
    if code.startswith("920"):
        return "北交"
    return "主板"


def _em_secid(code):
    """东财 secid：沪市(6/9开头)=1.，其余=0.。"""
    return ("1." if code.startswith(("6", "9")) else "0.") + code


def _norm_industry(ind):
    """行业名规范化：去掉申万罗马数字后缀（白酒Ⅱ→白酒）。"""
    if not ind:
        return None
    return re.sub(r"[ⅠⅡⅢⅣⅤ]+$", "", str(ind)).strip()


def fetch_industry_map(codes, cache_path=INDUSTRY_PATH):
    """东财批量行业映射（f100），增量缓存到 data/industry_map.json。
    数据源 push2delay.eastmoney.com（沙箱可达）；失败返回已有缓存。"""
    cache = load_json(cache_path, {})
    todo = [c for c in codes if c not in cache]
    if todo:
        for i in range(0, len(todo), 80):
            batch = todo[i:i + 80]
            secids = ",".join(_em_secid(c) for c in batch)
            url = ("https://push2delay.eastmoney.com/api/qt/ulist.np/get"
                   "?fltt=2&secids=%s&fields=f12,f14,f100" % secids)
            try:
                req = urllib.request.Request(url, headers={
                    "User-Agent": "Mozilla/5.0", "Referer": "https://quote.eastmoney.com/"})
                with urllib.request.urlopen(req, timeout=15) as r:
                    b = json.loads(r.read().decode("utf-8", "ignore"))
                got = 0
                for x in (b.get("data") or {}).get("diff") or []:
                    ind = x.get("f100")
                    if ind:
                        cache[str(x.get("f12"))] = ind
                        got += 1
                print("  [行业] 批次 %d-%d 命中 %d/%d" % (i, i + len(batch), got, len(batch)))
            except Exception as e:
                print("  [行业] 批量拉取失败: %s" % e)
        save_json(cache_path, cache)
    return cache


STYLE_ORDER = ["蓝筹", "科技成长", "主板", "北交"]


def build_flow(records, stocks, config):
    """资金流向 / 风格切换分析：
       - 每日 DK 背离趋势（D/K 净信号、当日均涨跌）
       - 风格/板块强弱（蓝筹 / 科技成长 / 主板 / 北交）按净信号排序 + 较昨日变化
       - 蓝筹 DK 轨迹与避险信号（当前 D 点 = 防御性买入信号）
       - 规则驱动的资金流向结论
    结论全由每日 DK 信号推断，仅供监测参考。"""
    bluechips = set(load_json(BLUECHIPS_PATH, {}).keys())
    days = records.get("days", [])
    if not days:
        return {"updated": records.get("updated"), "timeline": [], "styles": {},
                "style_order": STYLE_ORDER, "bluechips": [], "risk_signals": [], "conclusion": ""}
    names = stocks if isinstance(stocks, dict) else {}

    timeline = []
    style_hist = {s: [] for s in STYLE_ORDER}
    blue_traj = {}

    for d in days:
        date = d["date"]
        kset = set(d.get("k", []))
        dset = set(d.get("d", []))
        metrics = d.get("metrics", {})
        flagged = kset | dset
        allchg = [metrics.get(c) for c in flagged if isinstance(metrics.get(c), (int, float))]
        timeline.append({
            "date": date,
            "k": len(kset), "d": len(dset),
            "net": len(kset) - len(dset),                       # >0 卖出占优(净流出), <0 买入占优(净流入)
            "ratio": round(len(dset) / len(kset), 2) if kset else None,
            "avg_chg": round(sum(allchg) / len(allchg), 2) if allchg else None,
            "up": sum(1 for x in allchg if x > 0),
            "down": sum(1 for x in allchg if x < 0),
        })
        for s in STYLE_ORDER:
            sk = [c for c in kset if style_of(c, bluechips) == s]
            sd = [c for c in dset if style_of(c, bluechips) == s]
            schg = [metrics.get(c) for c in (set(sk) | set(sd)) if isinstance(metrics.get(c), (int, float))]
            style_hist[s].append({
                "date": date,
                "k": len(sk), "d": len(sd),
                "net": len(sk) - len(sd),
                "avg_chg": round(sum(schg) / len(schg), 2) if schg else None,
                "up": sum(1 for x in schg if x > 0),
                "down": sum(1 for x in schg if x < 0),
                "n": len(sk) + len(sd),
            })
        for c in bluechips:
            st = "K" if c in kset else ("D" if c in dset else "-")
            blue_traj.setdefault(c, []).append({"date": date, "state": st})

    # 行业映射（蓝筹 + 最新日 K/D 名单；东财 f100 增量缓存）
    latest = days[-1]
    latest_date = latest["date"]
    kset_l = set(latest.get("k", []))
    dset_l = set(latest.get("d", []))
    metrics_l = latest.get("metrics", {})
    ind_map = fetch_industry_map(set(bluechips) | kset_l | dset_l)

    # 蓝筹轨迹分析
    blue_list = []
    for c in sorted(bluechips):
        traj = blue_traj.get(c, [])
        states = [t["state"] for t in traj]
        current = states[-1] if states else "-"
        k2d = any(states[i - 1] == "K" and states[i] == "D" for i in range(1, len(states)))
        blue_list.append({
            "code": c, "name": names.get(c, c),
            "industry": _norm_industry(ind_map.get(c)) or "未分类",
            "traj": traj, "current": current,
            "k2d": k2d, "risk_off": (current == "D"), "risk_on": (current == "K"),
        })
    blue_list.sort(key=lambda b: ({"D": 0, "K": 1, "-": 2}.get(b["current"], 3), b["code"]))

    # 风格最新日 + 较昨日
    styles_out = {}
    for s in STYLE_ORDER:
        hist = style_hist[s]
        if not hist:
            continue
        last = hist[-1]
        prev = hist[-2] if len(hist) >= 2 else None
        net_delta = (last["net"] - prev["net"]) if prev else None
        chg_delta = ((last["avg_chg"] - prev["avg_chg"])
                     if (prev and last["avg_chg"] is not None and prev["avg_chg"] is not None) else None)
        if last["net"] > 0:
            direction = "净流出"          # 卖出(K)信号多于买入(D)
        elif last["net"] < 0:
            direction = "净流入"          # 买入(D)信号多于卖出(K)
        else:
            direction = "均衡"
        styles_out[s] = {"latest": last, "prev": prev,
                         "net_delta": net_delta, "chg_delta": chg_delta, "direction": direction}
    styles_ranked = sorted(styles_out.items(), key=lambda kv: kv[1]["latest"]["net"])  # 净流入(净最负)在前

    # 行业维度：最新日 D/K 按行业汇总（蓝筹动向单列——银行蓝筹 vs 科技蓝筹含义不同）
    ind_agg = {}
    for c in kset_l | dset_l:
        ind = _norm_industry(ind_map.get(c)) or "未分类"
        agg = ind_agg.setdefault(ind, {"d": 0, "k": 0, "chgs": [], "blue_d": [], "blue_k": []})
        m = metrics_l.get(c)
        if isinstance(m, (int, float)):
            agg["chgs"].append(m)
        if c in dset_l:
            agg["d"] += 1
            if c in bluechips:
                agg["blue_d"].append(names.get(c, c))
        if c in kset_l:
            agg["k"] += 1
            if c in bluechips:
                agg["blue_k"].append(names.get(c, c))
    industries = []
    for ind, agg in ind_agg.items():
        chgs = agg["chgs"]
        industries.append({
            "industry": ind, "d": agg["d"], "k": agg["k"],
            "net": agg["k"] - agg["d"],
            "avg_chg": round(sum(chgs) / len(chgs), 2) if chgs else None,
            "n": agg["d"] + agg["k"],
            "blue_d": agg["blue_d"], "blue_k": agg["blue_k"],
        })
    industries.sort(key=lambda x: (-x["n"], x["net"]))  # 信号数量多者在前，其次按净信号

    # 风险/流向信号
    risk_signals = []
    blue_d = [b for b in blue_list if b["current"] == "D"]
    blue_k = [b for b in blue_list if b["current"] == "K"]
    if blue_d:
        risk_signals.append({"type": "蓝筹避险", "level": "high",
            "text": "%d 只蓝筹出现 D 点（%s）—— 资金转向防御/避险信号增强"
                    % (len(blue_d), "、".join(b["name"] for b in blue_d))})
    if blue_k:
        risk_signals.append({"type": "蓝筹撤退", "level": "mid",
            "text": "%d 只蓝筹出现 K 点（%s）" % (len(blue_k), "、".join(b["name"] for b in blue_k))})
    if styles_ranked:
        inflow_s, inflow_v = styles_ranked[0]
        outflow_s, outflow_v = styles_ranked[-1]
        if inflow_v["latest"]["net"] < 0:
            risk_signals.append({"type": "资金流入", "level": "info",
                "text": "资金净流入方向：%s（D-K=%d，当日均涨 %s%%）"
                        % (inflow_s, -inflow_v["latest"]["net"], inflow_v["latest"]["avg_chg"])})
        if outflow_v["latest"]["net"] > 0:
            risk_signals.append({"type": "资金流出", "level": "warn",
                "text": "资金净流出方向：%s（K-D=%d）" % (outflow_s, outflow_v["latest"]["net"])})

    # 结论文本
    t_last = timeline[-1]
    concl = []
    concl.append("截至 %s：全市场 D=%d / K=%d，净卖出信号 %d 只，市场整体%s。"
                 % (latest_date, t_last["d"], t_last["k"], t_last["net"],
                    "偏弱（卖出信号占优）" if t_last["net"] > 0
                    else ("偏强（买入信号占优）" if t_last["net"] < 0 else "均衡")))
    if styles_ranked:
        inflow_s, inflow_v = styles_ranked[0]
        outflow_s, outflow_v = styles_ranked[-1]
        if inflow_v["latest"]["net"] < 0:
            concl.append("资金净流入风格：%s（当日均涨 %s%%）。" % (inflow_s, inflow_v["latest"]["avg_chg"]))
        if outflow_v["latest"]["net"] > 0:
            concl.append("资金净流出风格：%s（当日均涨 %s%%）。" % (outflow_s, outflow_v["latest"]["avg_chg"]))
    if blue_d or blue_k:
        parts = []
        if blue_d:
            parts.append("%d 只蓝筹现 D 点（%s）" % (len(blue_d), "、".join(b["name"] for b in blue_d)))
        if blue_k:
            parts.append("%d 只蓝筹现 K 点（%s）" % (len(blue_k), "、".join(b["name"] for b in blue_k)))
        concl.append("蓝筹动向：" + "；".join(parts) + "。蓝筹密集出现 D 点通常意味着资金转向防御/避险。")
    conclusion = "".join(concl) + "（结论由每日 DK 信号推断，仅供参考，非投资建议。）"

    return {
        "updated": latest_date,
        "timeline": timeline,
        "styles": styles_out,
        "style_order": STYLE_ORDER,
        "bluechips": blue_list,
        "industries": industries,
        "risk_signals": risk_signals,
        "conclusion": conclusion,
    }


def _build_ai_prompt(flow, er, obs=None):
    """把 DK 信号面 + 真实资金面数据组织成给智谱的分析 prompt。"""
    t = flow.get("timeline", [])
    tl = "\n".join("  %s: D=%d K=%d 净信号%+d 当日均涨%s%%" % (
        x["date"], x["d"], x["k"], x["net"], x.get("avg_chg")) for x in t[-5:])
    st = flow.get("styles", {})
    style_lines = []
    for s, v in st.items():
        L = v.get("latest") or {}
        style_lines.append("  %s: D=%d K=%d 净%+d %s 当日均涨%s%%" % (
            s, L.get("d", 0), L.get("k", 0), L.get("net", 0), v.get("direction", ""),
            L.get("avg_chg")))
    inds = (flow.get("industries") or [])[:12]
    ind_lines = "\n".join("  %s: D=%d K=%d 净%+d 当日均涨%s%% 蓝筹D[%s] 蓝筹K[%s]" % (
        x["industry"], x["d"], x["k"], x["net"], x.get("avg_chg"),
        "、".join(x.get("blue_d") or []), "、".join(x.get("blue_k") or [])) for x in inds) \
        if inds else "  （行业数据不足）"
    blues = flow.get("bluechips", [])
    blue_lines = "、".join("%s(%s·%s)" % (b["name"], b.get("industry", ""), b["current"])
                           for b in blues if b.get("current") != "-") or "无"
    er_lines = []
    if er and er.get("available"):
        er_lines = [
            "行业主力净流入TOP: " + "、".join("%s%s亿" % (x["name"], x["net"]) for x in (er.get("industry_in") or [])[:5]),
            "行业主力净流出TOP: " + "、".join("%s%s亿" % (x["name"], x["net"]) for x in (er.get("industry_out") or [])[:5]),
            "个股净流入TOP: " + "、".join("%s%s亿" % (x["name"], x["net"]) for x in (er.get("stocks_in") or [])[:5]),
            "个股净流出TOP: " + "、".join("%s%s亿" % (x["name"], x["net"]) for x in (er.get("stocks_out") or [])[:5]),
            "ETF分组净流入(亿): " + str(er.get("etf", {}).get("byGroup", {})),
            "财报舆情: " + str(er.get("earnings", {})),
        ]
    # 观察池实况（价格动作硬数据，用于约束情绪判断）
    obs_block = "（无）"
    if obs and isinstance(obs, dict):
        s = obs.get("summary", {})
        if s:
            obs_block = (
                "观察池样本=%d（活跃%d/已流出%d/待同步%d）；可评估=%d 只中 上涨%d / 下跌%d / 平%d；"
                "胜率=%s%% 平均累计涨幅=%s%% 中位累计涨幅=%s%%"
                % (
                    s.get("total_observed", 0), s.get("active", 0), s.get("exited", 0),
                    s.get("pending", 0), s.get("measurable", 0), s.get("up", 0),
                    s.get("down", 0), s.get("flat", 0), s.get("win_rate"),
                    s.get("avg_ret"), s.get("median_ret"),
                )
            )
    return """请基于以下数据输出一份A股资金流向与风格切换的市场解读（300-360字）。要求：
1. 先给出今日资金方向总体判断（流入/流出/避险）；
2. 指出风格与行业切换信号——哪些行业资金流入、哪些流出；特别区分蓝筹内部差异：银行等防御蓝筹出D点=避险，科技蓝筹出D点=进攻，含义不同；
3. 点评蓝筹避险信号与关键风险；
4. 若 DK 信号与真实主力资金方向矛盾，明确指出背离；
5. 【重点·必须以价格动作为准】结合【观察池实况】的涨跌家数与胜率、以及近几日「当日均涨」的正负变化，明确判断「市场情绪是否已修复 / 仍在退潮 / 刚刚企稳」——给出一句话结论。严格约束：(a) 只有当【观察池】胜率明显回升且当日均涨转正(大于0)时才可下「已修复」结论；(b) 若当日均涨为负或胜率仍低，不得写「均价涨转正回升」「已修复」，应如实写「仍在退潮/尚未修复/刚有企稳迹象」并交代依据；(c) 不要凭DK净信号计数自行脑补与价格相反的微观描述（如「D点持续减少」须对照时间线真实数字）。
语言流畅、口语化、像资深分析师给客户的盘中复盘，不要流水账列数据。

【DK信号面·每日趋势】
%s

【风格强弱】
%s

【行业强弱（按DK信号）】
%s

【蓝筹动向】%s

【观察池实况·价格动作（情绪判断的硬依据）】%s

【真实资金面（earnings-radar）】
%s""" % (tl, "\n".join(style_lines) or "  （数据不足）", ind_lines,
            blue_lines, obs_block,
            "\n".join(er_lines) if er_lines else "（数据缺失）")


def fetch_ai_summary(flow, er, config, obs=None, timeout=60):
    """调用智谱 GLM 生成资金流向文字解读；未配置 key 或失败返回 None（前端隐藏）。
    key 来源优先级：config.zhipu_api_key > 环境变量 ZHIPU_API_KEY > /root/.codebuddy/artifact/.zhipukey
    （key 只存沙箱，绝不进仓库/前端）。"""
    key = ""
    for src in (config.get("zhipu_api_key") if isinstance(config, dict) else None,
                os.environ.get("ZHIPU_API_KEY")):
        if src:
            key = str(src).strip()
            break
    if not key:
        try:
            p = "/root/.codebuddy/artifact/.zhipukey"
            if os.path.exists(p):
                key = open(p, encoding="utf-8").read().strip()
        except Exception:
            pass
    if not key:
        print("  [AI] 未配置智谱 key（config.zhipu_api_key / ZHIPU_API_KEY / .zhipukey），跳过文字解读")
        return None
    model = (config.get("zhipu_model") if isinstance(config, dict) else None) or "glm-4-flash"
    body = {
        "model": model,
        "messages": [
            {"role": "system", "content": "你是资深A股策略分析师，擅长通过资金流向与DK信号解读市场风格切换、板块轮动与风险。输出简洁专业的中文分析，直接给结论。"},
            {"role": "user", "content": _build_ai_prompt(flow, er, obs)},
        ],
        "temperature": 0.6,
        "max_tokens": 800,
    }
    req = urllib.request.Request(
        "https://open.bigmodel.cn/api/paas/v4/chat/completions",
        data=json.dumps(body).encode("utf-8"),
        headers={"Content-Type": "application/json", "Authorization": "Bearer " + key})
    try:
        with urllib.request.urlopen(req, timeout=timeout) as r:
            b = json.loads(r.read().decode("utf-8", "ignore"))
        text = ((b.get("choices") or [{}])[0].get("message") or {}).get("content") or ""
        text = text.strip()
        if not text:
            print("  [AI] 智谱返回空内容")
            return None
        print("  [AI] 智谱 %s 解读已生成（%d 字）" % (model, len(text)))
        return {"text": text, "model": model,
                "ts": datetime.datetime.now().strftime("%Y-%m-%d %H:%M")}
    except Exception as e:
        print("  [AI] 智谱调用失败: %s" % e)
        return None


def generate_app(records, stocks, obs, config):
    """生成手机端上传应用（含观察池、回踩K线预拉）。
    为彻底规避 GitHub 连接器对超长行/内容的截断与「手工转义引号」出错风险：
      - 种子数据 base64 切片为 seed_p1..N.js + seed_load.js
      - 整个 app HTML 也 base64 切片为 app_p1..M.js + app_load.js（app_load 用 document.write 注入）
      - 仓库里的 index.html 只做一个「壳」：顺序加载上述切片，自身几乎无引号、可安全推送
    所有切片文件均为 quote-light（base64），便于通过连接器逐文件可靠推送。"""
    import base64
    CHUNK = 5000

    def chunk_write(b64, prefix, var):
        parts = [b64[i:i + CHUNK] for i in range(0, len(b64), CHUNK)]
        for i, ch in enumerate(parts):
            line = ('window.%s = "%s";\n' % (var, ch)) if i == 0 \
                else ('window.%s += "%s";\n' % (var, ch))
            with open(os.path.join(BASE, "%s%d.js" % (prefix, i + 1)), "w", encoding="utf-8") as f:
                f.write(line)
        return len(parts)

    # 1) 种子数据切片（含预拉K线，供手机端无跨域使用）
    bluechips = sorted(set(load_json(BLUECHIPS_PATH, {}).keys()))
    klines = build_klines_for_pullback(records, days_window=30)
    flow = build_flow(records, stocks, config)
    er = build_realtime_er(stocks, config)
    flow["ai_summary"] = fetch_ai_summary(flow, er, config, obs)
    # 资金全景：个股→行业→大市主力资金（复用上面已拉取的个股实时数据，仅多 1 次批量请求）
    market = fetch_market_panorama(er.get("raw_stocks") or {}, records, obs)
    seed = {
        "updated": records.get("updated"),
        "days": records.get("days", []),
        "names": stocks,
        "bluechips": bluechips,
        "obs": obs,
        "klines": klines,
        "flow": flow,
        "er": er if er else {"available": False},
        "market": market,
        "em_pool": {"industry": EM_INDUSTRY_POOL,
                    "etf": [list(x) for x in EM_ETF_POOL]},
    }
    seed_b64 = base64.b64encode(json.dumps(seed, ensure_ascii=False).encode("utf-8")).decode("ascii")
    n = chunk_write(seed_b64, "seed_p", "__SEED_B64")
    with open(os.path.join(BASE, "seed_load.js"), "w", encoding="utf-8") as f:
        f.write('window.SEED_DATA = JSON.parse(new TextDecoder().decode('
                'Uint8Array.from(atob(window.__SEED_B64), c => c.charCodeAt(0))));\n')

    # 2) 整个 app HTML 切片（统一用绝对路径，根与 public 共用一套切片）
    tpl = open(APP_TEMPLATE_PATH, "r", encoding="utf-8").read()
    seed_tags = "".join('<script src="seed_p%d.js"></script>' % (i + 1) for i in range(n)) \
        + '<script src="seed_load.js"></script>'
    app_html = tpl.replace("__SEED__", "").replace("<!--SEED_SCRIPTS-->", seed_tags)
    app_b64 = base64.b64encode(app_html.encode("utf-8")).decode("ascii")
    m = chunk_write(app_b64, "app_p", "__APP_B64")
    with open(os.path.join(BASE, "app_load.js"), "w", encoding="utf-8") as f:
        f.write('document.open();document.write(new TextDecoder().decode('
                'Uint8Array.from(atob(window.__APP_B64), c => c.charCodeAt(0))));'
                'document.close();\n')

    # 3) 壳 index.html（quote-light）：只按顺序加载切片
    def shell(prefix=""):
        s = '<!DOCTYPE html>\n<html lang="zh-CN">\n<head>\n<meta charset="utf-8">\n' \
            '<title>DK 变化记录表</title>\n</head>\n<body>\n'
        for i in range(n):
            s += '<script src="%sseed_p%d.js"></script>\n' % (prefix, i + 1)
        s += '<script src="%sseed_load.js"></script>\n' % prefix
        for i in range(m):
            s += '<script src="%sapp_p%d.js"></script>\n' % (prefix, i + 1)
        s += '<script src="%sapp_load.js"></script>\n</body>\n</html>\n' % prefix
        return s
    for target, prefix in ((os.path.join(BASE, "index.html"), ""),
                           (os.path.join(BASE, "public", "index.html"), "../")):
        with open(target, "w", encoding="utf-8") as f:
            f.write(shell(prefix))

    # 清理旧的单文件
    for old in (os.path.join(BASE, "seed.js"), os.path.join(BASE, "public", "seed.js")):
        if os.path.exists(old):
            os.remove(old)
    print("  [生成] 壳 index.html + seed 切片 %d 个 + app 切片 %d 个 + 两个 loader" % (n, m))


def print_summary(analysis):
    print(
        "  [汇总] 更新=%s 跟踪=%d 今日K=%d 今日D=%d 反转个股=%d 累计反转=%d"
        % (
            analysis["updated"], analysis["total_stocks"], len(analysis["today_k"]),
            len(analysis["today_d"]), analysis["stocks_with_reversal"], analysis["total_reversals"],
        )
    )


# ---------- 自检 ----------
def run_selftest(config):
    stocks = {
        "600000": "浦发银行", "600001": "邯郸钢铁", "600002": "齐鲁石化",
        "600009": "上海机场", "600011": "华能国际",
    }
    days = [
        {"date": "2026-08-10", "k": ["600000", "600001", "600002"], "d": ["600009"], "metrics": {}},
        {"date": "2026-08-11", "k": ["600000", "600001"], "d": ["600002", "600009"], "metrics": {}},
        {"date": "2026-08-12", "k": ["600001"], "d": ["600000"], "metrics": {}},
        {"date": "2026-08-13", "k": ["600000"], "d": ["600001"], "metrics": {}},
        {"date": "2026-08-14", "k": [], "d": ["600000"], "metrics": {}},
    ]
    records = {"updated": "2026-08-14", "days": days}
    analysis = build_analysis(records, stocks, config)
    html = render_html(analysis, config)
    out = os.path.join(BASE, "_selftest.html")
    with open(out, "w", encoding="utf-8") as f:
        f.write(html)
    # 文本校验
    print("== 自检 ==")
    print_summary(analysis)
    for r in analysis["rows"]:
        if r["reversals"]:
            seq = " ".join(e["state"] for e in r["timeline"])
            print("  %s %s : %s  反转%d次 最近:%s->%s" % (
                r["code"], r["name"], seq, r["n_reversals"],
                r["latest_reversal"]["from"], r["latest_reversal"]["to"]))
    print("  自检网页已写: %s" % out)
    # 断言关键反转
    bycode = {r["code"]: r for r in analysis["rows"]}
    assert bycode["600002"]["n_reversals"] == 1, "600002 应 1 次反转 K->D"
    assert bycode["600000"]["n_reversals"] == 3, "600000 应 3 次反转"
    assert bycode["600001"]["n_reversals"] == 1, "600001 应 1 次反转 K->D"
    print("  断言通过 ✅")


# ---------- 入口 ----------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--k")
    ap.add_argument("--d")
    ap.add_argument("--date")
    ap.add_argument("--config", default=CONFIG_PATH)
    ap.add_argument("--regen", action="store_true")
    ap.add_argument("--selftest", action="store_true")
    args = ap.parse_args()

    config = load_config(args.config)

    if args.selftest:
        run_selftest(config)
        return

    if args.regen:
        records = load_json(RECORDS_PATH, {"updated": None, "days": []})
        purge_bj(records)
        stocks = load_json(STOCKS_PATH, {})
        analysis = build_analysis(records, stocks, config)
        obs = build_obs(records, config)
        generate_app(records, stocks, obs, config)
        print_summary(analysis)
        return

    if args.k and args.d:
        date = args.date or parse_date_from_filename(args.k, config) or parse_date_from_filename(args.d, config)
        if not date:
            raise SystemExit("[错误] 文件名无日期且未用 --date 指定，无法定位当天。")
        process_day(args.k, args.d, date, config)
        records = load_json(RECORDS_PATH, {"updated": None, "days": []})
        purge_bj(records)
        stocks = load_json(STOCKS_PATH, {})
        analysis = build_analysis(records, stocks, config)
        obs = build_obs(records, config)
        generate_app(records, stocks, obs, config)
        print_summary(analysis)
        return

    raise SystemExit("用法:\n  python process.py --k X_K.xlsx --d X_D.xlsx [--date 2026-08-14]\n  python process.py --regen\n  python process.py --selftest")


if __name__ == "__main__":
    main()
